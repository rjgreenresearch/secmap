#!/usr/bin/env python3
"""
afida_parser.py — Fixed for 2024 AFIDA Excel format

Parses USDA AFIDA detailed holdings data and matches entities against
SEC EDGAR CIK database. Identifies the visibility gap between entities
with federal SEC filings and private entities only visible through
state SOS records.

Fixes over original:
  - Auto-detects header row (AFIDA Excel files have title/grouped headers
    before the actual column names in Row 3)
  - Handles "Owner Name 1/" column naming
  - Handles "Number of Acres" column naming  
  - Extracts FIPS codes
  - Extracts Secondary Interest flags (China, Iran, Russia, North Korea)
  - Strips trailing whitespace from column names

Usage:
    python afida_parser.py --afida AFIDACurrentHoldingsYR2024.xlsx --tickers company_tickers.json --out output/

    # All adversarial countries:
    python afida_parser.py --afida AFIDA_2024.xlsx --tickers company_tickers.json --all-adversarial --out output/

Author: Robert J. Green | robert@rjgreenresearch.org
ORCID: 0009-0002-9097-1021
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import os
import re
import sys
from datetime import datetime
from difflib import SequenceMatcher
from typing import Dict, List, Optional, Set, Tuple

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column name aliases — covers all known AFIDA format variants
# ---------------------------------------------------------------------------

COLUMN_ALIASES = {
    "country": [
        "country", "country of investor", "investor_country", "foreign_country",
        "country_of_origin", "nation", "investor country", "foreign country",
    ],
    "entity_name": [
        "owner name 1/", "owner name", "owner name 1", "owner_name",
        "entity_name", "entity name", "investor_name", "investor name",
        "name", "owner", "beneficial_owner", "beneficial owner",
        "foreign person", "foreign_person", "holder", "filer",
    ],
    "state": [
        "state", "st", "state_code", "us_state", "location_state",
    ],
    "county": [
        "county", "county_name", "fips_county", "location_county",
    ],
    "fips": [
        "fips", "fips_code", "fips code", "county_fips", "county fips",
    ],
    "acres": [
        "number of acres", "number_of_acres", "acres", "acreage",
        "total_acres", "area_acres", "reported_acres",
    ],
    "owner_id": [
        "owner id", "owner_id", "ownerid",
    ],
    "parcel_id": [
        "parcel id", "parcel_id", "parcelid",
    ],
    "country_code": [
        "country code", "country_code", "countrycode",
    ],
    "owner_type": [
        "owner type", "owner_type", "ownertype",
    ],
    "type_of_interest": [
        "type of interest", "type_of_interest", "interest_type",
        "ownership_type", "interest",
    ],
    "percent_ownership": [
        "percent of ownership", "percent_of_ownership", "ownership_pct",
        "pct_ownership", "percent ownership",
    ],
    "acquisition_year": [
        "acquisition year", "acquisition_year", "acq_year", "year",
    ],
    "current_value": [
        "current value", "current_value", "value",
    ],
    "citizenship": [
        "citizenship",
    ],
    "secondary_china": [
        "secondary interest in china",
    ],
    "secondary_iran": [
        "secondary interest in iran",
    ],
    "secondary_russia": [
        "secondary interest in russia",
    ],
    "secondary_nk": [
        "secondary interest in north korea",
    ],
}

# Header detection keywords — if a row contains 3+ of these, it's the header
HEADER_SIGNATURES = {"state", "county", "country", "acres", "owner", "fips"}

# ---------------------------------------------------------------------------
# Country matching
# ---------------------------------------------------------------------------

CHINA_VARIANTS = {
    "china", "prc", "people's republic of china", "peoples republic of china",
    "p.r.c.", "chinese",
}

# Note: Hong Kong excluded from default China filter — it's a separate AFIDA
# country attribution. Use --include-hk to add it.
HONG_KONG_VARIANTS = {"hong kong", "hk", "macau", "macao"}

ADVERSARIAL_COUNTRIES = CHINA_VARIANTS | HONG_KONG_VARIANTS | {
    "russia", "russian federation",
    "iran", "islamic republic of iran",
    "north korea", "dprk", "democratic people's republic of korea",
    "belarus", "myanmar", "burma", "cuba", "venezuela", "syria", "nicaragua",
}


def normalize_country(raw: str) -> str:
    return raw.strip().lower().replace(".", "").replace("'", "'")


# ---------------------------------------------------------------------------
# Column resolution
# ---------------------------------------------------------------------------

def resolve_columns(headers: List[str]) -> Dict[str, str]:
    """
    Match actual column headers to expected field names using aliases.
    Strips whitespace and handles case-insensitive matching.
    """
    resolved = {}
    # Build lookup: normalized → original header name
    headers_clean = {}
    for h in headers:
        key = h.strip().lower().rstrip("/").rstrip()
        headers_clean[key] = h.strip()

    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            norm_alias = alias.strip().lower().rstrip("/").rstrip()
            if norm_alias in headers_clean:
                resolved[field] = headers_clean[norm_alias]
                break

    return resolved


# ---------------------------------------------------------------------------
# Entity name normalization and SEC matching
# ---------------------------------------------------------------------------

def normalize_entity_name(name: str) -> str:
    """Normalize entity name for matching."""
    n = name.upper().strip()
    # Remove common suffixes that vary between AFIDA and SEC
    for suffix in [", LLC", " LLC", ", LP", " LP", ", INC", " INC",
                   ", INC.", " INC.", ", CORP", " CORP", ", CORP.",
                   " CORP.", ", LTD", " LTD", ", LTD.", " LTD.",
                   ", CO.", " CO.", " CO", ", L.P.", " L.P.",
                   ", L.L.C.", " L.L.C."]:
        if n.endswith(suffix):
            n = n[:-len(suffix)]
    return n.strip()


def load_sec_tickers(path: str) -> Dict[str, dict]:
    """Load SEC company_tickers.json and build lookup structures."""
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    lookup = {}
    for key, entry in data.items():
        name = entry.get("title", "").strip()
        cik = entry.get("cik_str", "")
        ticker = entry.get("ticker", "")
        if name:
            normalized = normalize_entity_name(name)
            lookup[normalized] = {
                "name": name,
                "cik": str(cik),
                "ticker": ticker,
                "score": 1.0,
                "match_type": "exact",
            }
    logger.info("Loaded %d SEC tickers", len(lookup))
    return lookup


def match_entity_to_sec(entity_name: str, sec_lookup: Dict[str, dict],
                        threshold: float = 0.80) -> Optional[dict]:
    """Match an AFIDA entity name to SEC tickers."""
    normalized = normalize_entity_name(entity_name)

    # Exact match
    if normalized in sec_lookup:
        return sec_lookup[normalized]

    # Fuzzy match
    best_score = 0.0
    best_match = None
    for sec_name, sec_entry in sec_lookup.items():
        if abs(len(sec_name) - len(normalized)) > max(len(sec_name), len(normalized)) * 0.4:
            continue
        score = SequenceMatcher(None, normalized, sec_name).ratio()
        if score > best_score and score >= threshold:
            best_score = score
            best_match = {**sec_entry, "match_type": "fuzzy", "score": round(score, 3)}

    return best_match


# ---------------------------------------------------------------------------
# AFIDA loading — with auto header detection
# ---------------------------------------------------------------------------

def _detect_header_row(ws, max_scan: int = 10) -> int:
    """
    Scan the first max_scan rows to find the actual header row.
    The header row contains column names like 'State', 'County', 'Country'.
    AFIDA files often have title rows and grouped sub-headers above the
    actual column names.
    """
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        vals = {str(v or "").strip().lower() for v in row}
        matches = sum(1 for sig in HEADER_SIGNATURES if any(sig in v for v in vals))
        if matches >= 3:
            return row_idx
    return 1  # fallback


def load_afida_data(path: str) -> Tuple[List[str], List[dict]]:
    """
    Load AFIDA data from Excel (.xlsx, .xls) or CSV.
    Auto-detects header row for Excel files.
    Returns (headers, rows_as_dicts).
    """
    ext = os.path.splitext(path)[1].lower()

    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active

            # Auto-detect header row
            header_row = _detect_header_row(ws)
            logger.info("Detected header row: %d", header_row)

            # Read headers from the detected row
            headers = []
            for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
                headers = [str(h or "").strip() for h in row]

            # Read data starting from row after headers
            rows = []
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                d = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        d[headers[i]] = str(val or "").strip()
                # Skip completely empty rows
                if any(v for v in d.values()):
                    rows.append(d)

            wb.close()
            logger.info("Headers: %s", headers[:10])
            return headers, rows

        except ImportError:
            logger.error("openpyxl required for Excel files: pip install openpyxl")
            sys.exit(1)
    else:
        # CSV
        with open(path, "r", encoding="utf-8-sig") as f:
            sample = f.read(4096)
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t|;")
            f.seek(0)
            reader = csv.DictReader(f, dialect=dialect)
            headers = reader.fieldnames or []
            rows = list(reader)
        return headers, rows


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def run(
    afida_path: str,
    tickers_path: str,
    output_dir: str,
    countries: Optional[Set[str]] = None,
    match_threshold: float = 0.80,
    china_only: bool = True,
    include_hk: bool = False,
    include_secondary: bool = True,
):
    """Parse AFIDA data, match against SEC tickers, produce output files."""
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    # 1. Load data
    logger.info("Loading AFIDA data from %s", afida_path)
    headers, all_rows = load_afida_data(afida_path)
    logger.info("Loaded %d rows, %d columns", len(all_rows), len(headers))

    # 2. Resolve columns
    col_map = resolve_columns(headers)
    logger.info("Resolved columns: %s", {k: v for k, v in col_map.items()})

    # Check required columns
    missing = [f for f in ["country", "entity_name"] if f not in col_map]
    if missing:
        logger.error(
            "Could not resolve required columns: %s\n"
            "Available headers: %s\n"
            "Edit COLUMN_ALIASES to match your file.",
            missing, headers
        )
        sys.exit(1)

    # 3. Build country filter
    if countries:
        filter_set = countries
    elif china_only:
        filter_set = set(CHINA_VARIANTS)
        if include_hk:
            filter_set |= HONG_KONG_VARIANTS
    else:
        filter_set = ADVERSARIAL_COUNTRIES

    # 4. Filter rows — by country field OR secondary interest flags
    filtered = []
    secondary_col = col_map.get("secondary_china")

    for row in all_rows:
        country_raw = row.get(col_map["country"], "")
        country_match = normalize_country(country_raw) in filter_set

        # Also check Secondary Interest in China flag
        secondary_match = False
        if include_secondary and secondary_col:
            sec_val = row.get(secondary_col, "").strip()
            if sec_val == "1":
                secondary_match = True

        if country_match or secondary_match:
            filtered.append(row)

    logger.info(
        "Filtered to %d rows (country match + secondary interest)",
        len(filtered),
    )

    if not filtered:
        logger.warning("No rows matched. Check column mapping and country filter.")
        return

    # 5. Extract unique entities with enriched data
    name_col = col_map["entity_name"]
    entities = {}
    for row in filtered:
        name = row.get(name_col, "").strip()
        if not name:
            continue
        if name not in entities:
            entities[name] = {
                "entity_name": name,
                "country": row.get(col_map.get("country", ""), ""),
                "country_code": row.get(col_map.get("country_code", ""), ""),
                "state": row.get(col_map.get("state", ""), ""),
                "counties": set(),
                "fips_codes": set(),
                "total_acres": 0.0,
                "holding_count": 0,
                "owner_type": row.get(col_map.get("owner_type", ""), ""),
                "secondary_china": False,
                "secondary_iran": False,
                "secondary_russia": False,
                "secondary_nk": False,
            }
        e = entities[name]

        # County
        county = row.get(col_map.get("county", ""), "")
        if county:
            e["counties"].add(county)

        # FIPS
        fips = row.get(col_map.get("fips", ""), "")
        if fips:
            e["fips_codes"].add(str(fips).split(".")[0].zfill(5))

        # Acres
        try:
            acres_str = row.get(col_map.get("acres", ""), "0")
            acres_str = re.sub(r"[^0-9.]", "", str(acres_str))
            e["total_acres"] += float(acres_str) if acres_str else 0
        except (ValueError, TypeError):
            pass

        e["holding_count"] += 1

        # Secondary interest flags
        for flag_field, flag_key in [
            ("secondary_china", "secondary_china"),
            ("secondary_iran", "secondary_iran"),
            ("secondary_russia", "secondary_russia"),
            ("secondary_nk", "secondary_nk"),
        ]:
            col = col_map.get(flag_field)
            if col and row.get(col, "").strip() == "1":
                e[flag_key] = True

    logger.info("Unique entities: %d", len(entities))

    # 6. Match against SEC
    sec_lookup = load_sec_tickers(tickers_path)

    matched = []
    unmatched = []

    for name, info in sorted(entities.items()):
        match = match_entity_to_sec(name, sec_lookup, match_threshold)
        # Convert sets to strings for CSV
        info["counties"] = "; ".join(sorted(info["counties"]))
        info["fips_codes"] = "; ".join(sorted(info["fips_codes"]))
        info["total_acres"] = round(info["total_acres"], 2)

        if match:
            matched.append({
                **info,
                "cik": match["cik"],
                "sec_name": match["name"],
                "ticker": match["ticker"],
                "match_type": match["match_type"],
                "match_score": match["score"],
            })
        else:
            unmatched.append({
                **info,
                "cik": "",
                "sec_name": "",
                "ticker": "",
                "match_type": "none",
                "match_score": 0.0,
            })

    # 7. Write outputs
    matched_path = os.path.join(output_dir, f"afida_sec_matched.csv")
    unmatched_path = os.path.join(output_dir, f"afida_unmatched.csv")
    summary_path = os.path.join(output_dir, f"afida_parse_summary.txt")
    cik_list_path = os.path.join(output_dir, f"secmap_target_ciks.txt")

    fieldnames = [
        "entity_name", "country", "country_code", "state", "counties",
        "fips_codes", "total_acres", "holding_count", "owner_type",
        "secondary_china", "secondary_iran", "secondary_russia", "secondary_nk",
        "cik", "sec_name", "ticker", "match_type", "match_score",
    ]

    with open(matched_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        writer.writerows(matched)

    with open(unmatched_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        writer.writerows(unmatched)

    # CIK list for SECMap
    ciks = sorted(set(m["cik"] for m in matched if m["cik"]))
    with open(cik_list_path, "w", encoding="utf-8") as f:
        f.write("# SECMap target CIKs from AFIDA cross-reference\n")
        f.write(f"# Generated: {timestamp}\n")
        f.write(f"# Source: {os.path.basename(afida_path)}\n")
        f.write(f"# Matched {len(matched)} of {len(entities)} entities\n")
        f.write(f"# Unique CIKs: {len(ciks)}\n\n")
        for cik in ciks:
            names = [m["sec_name"] for m in matched if m["cik"] == cik]
            f.write(f"{cik}  # {names[0]}\n")

    # Summary
    total = len(entities)
    matched_count = len(matched)
    unmatched_count = len(unmatched)
    coverage_pct = (matched_count / total * 100) if total else 0
    gap_pct = 100 - coverage_pct

    total_acres = sum(e["total_acres"] for e in entities.values())
    matched_acres = sum(m["total_acres"] for m in matched)
    unmatched_acres = sum(u["total_acres"] for u in unmatched)

    summary_lines = [
        "=" * 70,
        "AFIDA PARSER — COVERAGE ANALYSIS",
        "=" * 70,
        f"Source file:          {os.path.basename(afida_path)}",
        f"Generated:           {timestamp}",
        f"Total AFIDA rows:    {len(all_rows)}",
        f"Filtered rows:       {len(filtered)}",
        f"Country filter:      {', '.join(sorted(filter_set))}",
        "",
        "-" * 70,
        "ENTITY SUMMARY",
        "-" * 70,
        f"Unique entities:     {total}",
        f"Total acreage:       {total_acres:,.1f} acres",
        f"Unique counties:     {len(set().union(*(e['fips_codes'] for e in entities.values() if isinstance(e['fips_codes'], set) or True)))}",
        "",
        "-" * 70,
        "SEC COVERAGE GAP — THE ARTICLE 2 FINDING",
        "-" * 70,
        f"SEC-matched entities:   {matched_count:>4} ({coverage_pct:.1f}%)",
        f"SEC-matched acreage:    {matched_acres:>12,.1f} acres",
        f"Unmatched entities:     {unmatched_count:>4} ({gap_pct:.1f}%) ← INVISIBLE TO FEDERAL OWNERSHIP ANALYSIS",
        f"Unmatched acreage:      {unmatched_acres:>12,.1f} acres",
        "",
        f"Unique CIKs for SECMap: {len(ciks)}",
        "",
    ]

    if ciks:
        summary_lines.append("-" * 70)
        summary_lines.append("SEC-MATCHED ENTITIES (feed to SECMap)")
        summary_lines.append("-" * 70)
        for m in sorted(matched, key=lambda x: -x["total_acres"]):
            summary_lines.append(
                f"  CIK {m['cik']:>10}  {m['entity_name'][:40]:<40}  "
                f"{m['total_acres']:>10,.1f} ac  ({m['match_type']}, {m['match_score']})"
            )

    summary_lines.append("")
    summary_lines.append("-" * 70)
    summary_lines.append("TOP UNMATCHED ENTITIES — STATE SOS TARGETS")
    summary_lines.append("-" * 70)
    for u in sorted(unmatched, key=lambda x: -x["total_acres"])[:30]:
        flags = []
        if u["secondary_china"]: flags.append("SEC_CHINA")
        if u["secondary_iran"]: flags.append("SEC_IRAN")
        if u["secondary_russia"]: flags.append("SEC_RUSSIA")
        flag_str = f"  [{','.join(flags)}]" if flags else ""
        summary_lines.append(
            f"  {u['entity_name'][:45]:<45}  {u['total_acres']:>10,.1f} ac  "
            f"{u['state'][:2]:>2}  {u['country'][:15]}{flag_str}"
        )

    summary_lines.extend([
        "",
        "=" * 70,
        "OUTPUT FILES",
        "=" * 70,
        f"  Matched:    {matched_path}",
        f"  Unmatched:  {unmatched_path}",
        f"  CIK list:   {cik_list_path}",
        f"  Summary:    {summary_path}",
        "",
        "NEXT STEPS:",
        f"  1. Add CIKs from {os.path.basename(cik_list_path)} to SECMap TARGET_CIKS",
        f"  2. Run SECMap production batch: python run_production.py",
        f"  3. Research unmatched entities via state SOS records",
        f"  4. Feed state SOS results into GapAnalyzer",
    ])

    summary_text = "\n".join(summary_lines)
    with open(summary_path, "w", encoding="utf-8") as f:
        f.write(summary_text)

    # Print summary to console
    print(summary_text)

    logger.info("Done. Outputs in %s", output_dir)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="AFIDA Parser — Cross-reference AFIDA holdings with SEC EDGAR",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s --afida AFIDA_2024.xlsx --tickers company_tickers.json --out output/
  %(prog)s --afida AFIDA_2024.xlsx --tickers company_tickers.json --all-adversarial --out output/
  %(prog)s --afida AFIDA_2024.xlsx --tickers company_tickers.json --include-hk --out output/
        """
    )
    parser.add_argument("--afida", required=True, help="Path to AFIDA Excel or CSV file")
    parser.add_argument("--tickers", required=True, help="Path to SEC company_tickers.json")
    parser.add_argument("--out", default="output", help="Output directory (default: output/)")
    parser.add_argument("--all-adversarial", action="store_true",
                        help="Include all adversarial countries (default: China only)")
    parser.add_argument("--include-hk", action="store_true",
                        help="Include Hong Kong/Macau in China filter")
    parser.add_argument("--threshold", type=float, default=0.80,
                        help="Fuzzy match threshold (default: 0.80)")
    parser.add_argument("--no-secondary", action="store_true",
                        help="Don't include Secondary Interest flag matches")

    args = parser.parse_args()

    run(
        afida_path=args.afida,
        tickers_path=args.tickers,
        output_dir=args.out,
        china_only=not args.all_adversarial,
        include_hk=args.include_hk,
        match_threshold=args.threshold,
        include_secondary=not args.no_secondary,
    )


if __name__ == "__main__":
    main()
